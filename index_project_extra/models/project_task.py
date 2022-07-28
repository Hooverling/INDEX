# -*- coding: utf-8 -*-
import logging
from odoo import models, fields, api
from odoo.exceptions import ValidationError, UserError, RedirectWarning
from tempfile import NamedTemporaryFile
from datetime import datetime, date
from odoo.tools.translate import _
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from io import BytesIO 
import base64
_logger = logging.getLogger(__name__)

class Tasks(models.Model):
    _inherit = "project.task"
    custom_task_line_ids = fields.One2many(comodel_name='custom.task.line', inverse_name='task_id')
    valida_ids = fields.One2many(comodel_name='custom.vlines', inverse_name='v_id')
    validated = fields.Boolean('Validación Index')
    email_sent = fields.Boolean('Correo Enviado')
    stage_seq = fields.Integer(string='Secuencia', related='stage_id.sequence')
    #Index
    vidate = fields.Datetime('Validación index Fecha')
    viuser_id = fields.Many2one('res.users', string='Validación Index usuario')
    #Forwarder
    vfdate = fields.Datetime('Validación Forwarder Fecha')
    vfuser_id = fields.Many2one('res.users', string='Validación Forwarder usuario')
    vfdocs = fields.Boolean('Cumplimiento Documentación')
    vflete = fields.Boolean('Cumplimiento Costo Flete Marítimo')
    #Agente Aduanal
    vadate = fields.Datetime('Validación Agente Aduanal Fecha')
    vauser_id = fields.Many2one('res.users', string='Validación Agente Aduanal usuario')
    varevalida =    fields.Boolean('Revalidación del BL')
    vafolio =       fields.Boolean('Liberación del Folio')    
    vaprevio =      fields.Boolean('Programación del Previo')
    vamaniobra =    fields.Boolean('Maniobra de Carga')
    #Transportista
    vtdate =    fields.Datetime('Validación Transportista Fecha')
    vtuser_id = fields.Many2one('res.users', string='Validación Transportista usuario')
    vtplaca =       fields.Char('Nº Placa')
    vtconductor =   fields.Char('Nombre del Conductor')
    
    
    def userfrompartner(self,p_id):
        if p_id == False:
            return False
        usr_id= self.env['res.users'].search([('partner_id','=',p_id.id)],limit = 1)
        if usr_id:#si se encontró el usuario lo regresa
            _logger.error('para el partner '+str(p_id.name)+' Se encontró el Usuario '+str(usr_id.name))
            return usr_id
        return False #no lo encontramos regresamos Error

    
    #Validacion Index
    def val_index(self):
        #Validamos que las horas entre el ETA y lahora actual esten entre 48 y 72 horas
        #obtenemos el eta
        if self.email_sent == False:
            raise ValidationError('No se ha enviado el Correo')
        if len(self.custom_task_line_ids) == 0:
            raise ValidationError('El reporte esta vacio')
        eta = self.custom_task_line_ids[0].eta_date
        if eta == False:
            raise ValidationError('No se ha asignado una fecha Eta en el Reporte')
        dif =  eta - datetime.now()
        horas = int(dif.total_seconds()/3600)
        if horas < 48:
            self.kanban_state = "blocked"
            body = '->Hay '+str(horas)+' horas de diferencia solo se permiten permiten diferencias mayores a  48 horas entre la la fecha actual y la estimada la tarea será Bloqueada'
            self.message_post(body=body)
            return self
            
        user = self.env['res.users'].browse(self._context.get('uid'))
        self.viuser_id = user
        self.vidate= datetime.now()
        body = str(self.vidate) +"->Se ha autorizado el pase a la etapa En Proceso "
        self.message_post(body=body)
        #Eliminamos posibles ciclos anteriores
        for delval in self.valida_ids:
            if delval:
                delval.unlink()
        vlines = self.env['custom.vlines']
        #Generamos el ciclo de validac  
        for cv in self.custom_task_line_ids:
            f_user = self.userfrompartner(cv.forwarders)
            a_user = self.userfrompartner(cv.agente_aduanal)
            t_user = self.userfrompartner(cv.transportista)
            values = {
                'v_id'              : self.id,
                'container_number'  : cv.container_number,
                'eta_date'          : cv.eta_date,
                'custom_category'   : cv.custom_category,
                'etapa'             :  '1',
                'u_forwarder'         : f_user.id,
                'u_aduanal'           : a_user.id,
                'u_transportista'     : t_user.id,
            }
            vlines.create(values)
        #Mandamos un correo a los contactos de la etapa
        #ahora vamos por la lista de correos de la etapa
        l_mails =[]
        for lm in self.stage_id.emails:
                if lm.email:
                    l_mails.append(lm.email)
        #seguramente hay duplicados vamos a eliminarlos
        l_mails = list(dict.fromkeys(l_mails))
        emto = ''
        for lm in l_mails:
            emto = emto + str(lm) + ','
        #Cuerpo del correo
        body_mail = 'Buen día' +'\n'+'El proceso '+str(self.name)+' ha iniciado.'
        body_mail_html = '<p>Buen d&iacute;a.</p><p> El Proceso '+str(self.name)+' ha iniciado.</p>'
        #ya tenemos todo mandemos el correo
        mail_pool = self.env['mail.mail']
        values={}
        values.update({'subject': 'Proceso Iniciado'})
        values.update({'email_to': emto})
        values.update({'body_html': body_mail_html })
        values.update({'body': body_mail })
        #values.update({'attachment_ids': inserted_id })
        values.update({'res_id': self.id }) #[optional] here is the record id, where you want to post that email after sending
        values.update({'model': 'project.task' }) #[optional] here is the object(like 'project.project')  to whose record id you want to post that email after sending
        msg_id = mail_pool.create(values)
        if msg_id:
            mail_pool.send([msg_id])                    
        #Buscamos el siguiente stage en la secuencia
        for st in self.project_id.type_ids:
            if st.sequence == 1:#es la siguiente secuencia
                self.stage_id = st #cambiamos el stage
        return self    

    #reset de la Tarea regresa al stage inicial
    def reset_index(self):
        self.active = 1
        self.kanban_state = "blocked"
        self.email_sent = 0
        #Eliminamos posibles ciclos anteriores
        for delval in self.valida_ids:
            if delval:
                delval.unlink()        
        #Buscamos el siguiente stage inicial en la secuencia
        for st in self.project_id.type_ids:
            if st.sequence == 0:#es la siguiente secuencia
                self.stage_id = st #cambiamos el stage
        body = "->Se ha regresado la tarea a la etapa Inicial "
        self.message_post(body=body)
        return self
    #Alta (en caso de baja por error)
    def alta_index(self):
        self.active = 1
        body = "->Se ha regresado la tarea como activa "
        self.message_post(body=body)
        return self

    #Baja por Index
    def baja_index(self):
        self.active = 0
        body = "->Se ha marcado la tarea como archivada (Baja)"
        self.message_post(body=body)
        return self

    def fin_index(self):
        for i in self.valida_ids:
            #raise ValidationError(str(i.etapa))
            if i.etapa not in ('0','5'):
                raise ValidationError('Solo se puede finalizar la Tarea si todas las validaciones estan en Estapa Finalizada o Baja')
        #Buscamos la primer etapa de la secuencia
        for st in self.project_id.type_ids:
            if st.sequence == 2:#es la siguiente secuencia
                self.stage_id = st #cambiamos el stage
        #Mandamos un correo a los contactos de la etapa (final)
        #ahora vamos por la lista de correos de la etapa
        l_mails =[]
        for lm in self.stage_id.emails:
                if lm.email:
                    l_mails.append(lm.email)
        #seguramente hay duplicados vamos a eliminarlos
        l_mails = list(dict.fromkeys(l_mails))
        emto = ''
        for lm in l_mails:
            emto = emto + str(lm) + ','
        #Cuerpo del correo
        body_mail = 'Buen día' +'\n'+'El proceso '+str(self.name)+' ha finalizado.'
        body_mail_html = '<p>Buen d&iacute;a.</p><p> El Proceso '+str(self.name)+' ha finalizado.</p>'
        #ya tenemos todo mandemos el correo
        mail_pool = self.env['mail.mail']
        values={}
        values.update({'subject': 'Proceso Finalizado'})
        values.update({'email_to': emto})
        values.update({'body_html': body_mail_html })
        values.update({'body': body_mail })
        #values.update({'attachment_ids': inserted_id })
        values.update({'res_id': self.id }) #[optional] here is the record id, where you want to post that email after sending
        values.update({'model': 'project.task' }) #[optional] here is the object(like 'project.project')  to whose record id you want to post that email after sending
        msg_id = mail_pool.create(values)
        if msg_id:
            mail_pool.send([msg_id])                      
        return self

    def send_email(self):
        #validamos que todas las fechas eta seanlas mismas asi como las categorías
        if len (self.custom_task_line_ids) == 0:
            raise ValidationError('El reporte esta vacío')
        eta0 = self.custom_task_line_ids[0].eta_date
        #validamos que el eta no tenga mas de 48 horas de diferencia
        if eta0:
            dif =  eta0 - datetime.now()
            horas = int(dif.total_seconds()/3600)
            if horas < 48:
                raise ValidationError('Hay '+str(horas)+' horas de diferencia solo se permiten diferencias mayores a 48 horas entre la la fecha actual y la estimada' )

        cat0 =   self.custom_task_line_ids[0].custom_category
        oper0 =  self.custom_task_line_ids[0].operadora
        buque0 = self.custom_task_line_ids[0].buque

        #corremos ciclo de validaciones de integridad de Información
        for v in self.custom_task_line_ids:
            if v.eta_date != eta0: #todas las fechas Eta deben ser iguales
                raise ValidationError('La línea con Contenedor '+str(v.container_number)+' contiene una fecha ETA diferente a la especificada en la primer línea')
            if v.custom_category != cat0:#todas las Categorías deben ser iguales
                raise ValidationError('La línea con Contenedor '+str(v.container_number)+' contiene una categoría diferente a la especificada en la primer línea')
            if v.operadora != oper0:#todas las Operadoras deben ser iguales
                raise ValidationError('La línea con Contenedor '+str(v.container_number)+' contiene una Operadora diferente a la especificada en la primer línea')
            if v.buque != buque0:#todas los Buques deben ser iguales
                raise ValidationError('La línea con Contenedor '+str(v.container_number)+' contiene un buque diferente al especificado en la primer línea')
            if v.eta_date == False or v.dispatch_date == False:
                raise ValidationError('La línea con Contenedor '+str(v.container_number)+' contiene Fechas Estimada o de Despacho vacias')
            dif = v.dispatch_date -v.eta_date
            horas = int(dif.total_seconds()/3600)
            if horas <= 24:
                raise ValidationError('La línea con Contenedor '+str(v.container_number)+' La fecha de despacho debe ser mayor a la estimada por al menos 24 horas')
            if self.userfrompartner(v.forwarders) == False:
                raise ValidationError('La línea con Contenedor '+str(v.container_number)+' Especifica al Forwarder '+str(v.forwarders.name)+' Pero este no posee un usuario relacionado')
            if self.userfrompartner(v.agente_aduanal) == False:
                raise ValidationError('La línea con Contenedor '+str(v.container_number)+' Especifica al Agente Aduanal '+str(v.agente_aduanal.name)+' Pero este no posee un usuario relacionado')
            if self.userfrompartner(v.transportista) == False:
                raise ValidationError('La línea con Contenedor '+str(v.container_number)+' Especifica al Transportista '+str(v.transportista.name)+' Pero este no posee un usuario relacionado')

        #Iniciamos con la cabecera del Excel
        r_type = '24' #tipo de reporte
        nav ='NO Asignado' #naviera
        buque ='NO Asignado' #buque
        operadora ='NO Asignado' #operadora
        cat = ''
        wb = Workbook() #creamos objeto
        ws = wb.active # inicializamos
        reng = 9 #indicador de renglon
        ws.title = "Solicitud" #titulo
        ws.cell(7, 3).value = "Solicitud de Marca de Calidad IMMEX"
        ws.cell(7, 3).font = Font(size = "15")
        #--------------------Cabecera------------------------------------
        ws['A8'] = 'CATEGORIA'
        ws['B8'] = 'BL'
        ws['C8'] = '#CONTENEDOR'
        ws['D8'] = 'TIPO DE CONTENEDOR'
        ws['E8'] = 'ID AGENTE ADUANAL'
        ws['F8'] = 'ID NAVIERA'
        ws['G8'] = 'ID FORWARDERS'
        ws['H8'] = 'OPERADORA'
        ws['I8'] = 'BUQUE'
        ws['J8'] = 'NO. VIAJE'
        ws['K8'] = 'FECHA DE ETA'
        ws['L8'] = 'FECHA PREVIO'
        ws['M8'] = 'FECHA DESPACHO'
        ws['N8'] = 'PREVIO'
        ws['O8'] = 'PESO'
        ws['P8'] = 'PIEZAS'
        ws['Q8'] = 'EMBALAJE'
        

        for col in range (1,18):
            ws.cell(row=8, column=col).font = Font(color="FFFFFF")
            ws.cell(row=8, column=col).fill = PatternFill('solid', fgColor = '063970')
        #Traemos Imagen del logo
        if self.stage_id.attach_document:
            buf_image= BytesIO(base64.b64decode(self.stage_id.attach_document))
            img = Image(buf_image)
            img.anchor='A1'
            ws.add_image(img)
            
            
        #Traemos todos los clientes activos 
        for i in self.custom_task_line_ids:
            if str(i.custom_category) == '24':
                cat = 'IMMEX 24 hrs'
            if str(i.custom_category) == '36':
                cat = 'IMMEX 36 hrs'
            ws.cell(row=reng, column=1).value = cat #i.custom_category
            ws.cell(row=reng, column=2).value = i.bl
            ws.cell(row=reng, column=3).value = i.container_number
            if i.container_type_id.code:
                ws.cell(row=reng, column=4).value = str(i.container_type_id.code)
            ws.cell(row=reng, column=5).value = i.agente_aduanal.name
            ws.cell(row=reng, column=6).value = i.naviera
            ws.cell(row=reng, column=7).value = i.forwarders.name
            ws.cell(row=reng, column=8).value = i.operadora.name
            ws.cell(row=reng, column=9).value = i.buque
            ws.cell(row=reng, column=10).value = i.numero_viaje
            ws.cell(row=reng, column=11).value = i.eta_date
            if i.previo_date:
                ws.cell(row=reng, column=12).value = i.previo_date
            if i.dispatch_date:
                ws.cell(row=reng, column=13).value = str(i.dispatch_date)
            if i.service_type_id.code:                
                ws.cell(row=reng, column=14).value = str(i.service_type_id.code)
            if i.peso:                
                ws.cell(row=reng, column=15).value = i.peso
            if i.pieza:
                ws.cell(row=reng, column=16).value = i.pieza
            if i.packing_type_id.code:
                ws.cell(row=reng, column=17).value = str(i.packing_type_id.code)
            reng = reng + 1
        with NamedTemporaryFile() as tmp: #graba archivo temporal
            wb.save(tmp.name) #graba el contenido del excel en tmp.name
            output = tmp.read()
        #filename = 'Solicitud De Marca%s.xlsx' % (date.today().strftime('%Y%m%d')) #nombre del archivo en Excel
        #filename =  self.name +'-%s.xlsx' % (date.today().strftime('%Y%m%d')) #nombre del archivo en Excel'
        filename = 'SOLICITUD MC IMMEX ' +str(self.partner_id.name)+' // '+str(cat0)+'hrs// '+str(oper0.name)+' //BUQUE '+str(buque0)+'//VIAJE '+str(i.numero_viaje)+'//CONTENEDORES '+str(len(self.custom_task_line_ids))
        self.name = filename
        filename = filename+'.xlsx'
        #raise ValidationError(str(filename))
        xlsx = {                            #características del archivo
                'name': filename,
                'type': 'binary',
                'res_model': 'selmrp.tmpexploit',
                'datas': base64.b64encode(output),  #aqui metemos el archivo generado y grabado
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                }
        inserted_id=self.env['ir.attachment'].create(xlsx) # creamos el link de download
        url='/web/content/%s?download=1' %(inserted_id.id)
        #obtenemos la lista de todos los emails a mandar
        #primero vamos por los que tiene el reporte 
        l_mails = []
        #Validación de Correos 
        for i in self.custom_task_line_ids:
            if i.forwarders.email == False or i.forwarders.email2 == False:
                raise ValidationError('El Forwarder '+str(i.forwarders.name)+' No tiene correos asignados se cancela la operación')
            if i.operadora.email == False:
                raise ValidationError('La Operadora  '+str(i.operadora.name)+' No tiene correos asignados se cancela la operación')
            if i.agente_aduanal.email == False:
                raise ValidationError('El Agente Aduanal'+str(i.agente_aduanal.name)+' No tiene correos asignados se cancela la operación')
        #Validación Esta ligado al Imex??? 
        l_forw =[] #forwarders admitidos
        l_aa = []  #Agentes aduanales admitidos
        l_oper = [] #operadores admitidos
        l_trans = [] #Transportistas admitidos
        for i in self.partner_id.partners_ids:
            if i.partner_type == 'F': #forwarders
                for j in i.partner:
                    l_forw.append(j.id) 
            if i.partner_type == 'A': #Agentes Aduanales
                for j in i.partner:
                    l_aa.append(j.id) 
            if i.partner_type == 'O': #Operadoras
                for j in i.partner:
                    l_oper.append(j.id) 
            if i.partner_type == 'T': #Operadoras
                for j in i.partner:
                    l_trans.append(j.id) 
        if len(l_forw) == 0:
            raise ValidationError('Immex '+str(self.partner_id.name)+' no tiene Forwarders Relacionados')
        if len(l_aa) == 0:
            raise ValidationError('Immex '+str(self.partner_id.name)+' no tiene Agentes Aduanales Relacionados')            
        if len(l_oper) == 0:
            raise ValidationError('Immex '+str(self.partner_id.name)+' no tiene Operadores Relacionados')            
        if len(l_trans) == 0:
            raise ValidationError('Immex '+str(self.partner_id.name)+' no tiene Transportistas Relacionados')            

        for i in self.custom_task_line_ids:
            if i.forwarders.id not in l_forw:
                raise ValidationError('El Forwarder '+str(i.forwarders.name)+' No esta asignado al contacto IMMEX se cancela la operación')
            if i.operadora.id not in l_oper:
                raise ValidationError('La Operadora  '+str(i.operadora.name)+' No esta asignado al contacto IMMEX se cancela la operación')
            if i.agente_aduanal.id not in l_aa:
                raise ValidationError('El Agente Aduanal '+str(i.agente_aduanal.name)+' No esta asignado al contacto IMMEX  se cancela la operación')
            if i.transportista.id not in l_trans:
                raise ValidationError('El Transportista '+str(i.agente_aduanal.name)+' No esta asignado al contacto IMMEX  se cancela la operación')

            #si se trata de IMMEX 24 se agrega el email 1
            if i.custom_category == '24':
                r_type = '24'
                l_mails.append(i.forwarders.email)
                l_mails.append(i.operadora.email)
                l_mails.append(i.agente_aduanal.email)
            #si se trata de IMMEX 26 se agrega el email 2
            if i.custom_category == '36':
                r_type = '36'
                l_mails.append(i.forwarders.email2)
                l_mails.append(i.operadora.email2)
                l_mails.append(i.agente_aduanal.email2)
        #ahora vamos por la lista de correos de la etapa

        for lm in self.stage_id.emails:
                if lm.email:
                    l_mails.append(lm.email)
        #seguramente hay duplicados vamos a eliminarlos
        l_mails = list(dict.fromkeys(l_mails))
        emto = ''
        for lm in l_mails:
            emto = emto + str(lm) + ','
        #Cuerpo del correo
        body_mail = 'Buen día' +'\n'+'En relación al proceso de MC IMMEX, comparto el formato de solicitud para el proceso debido con categoría de '+str(r_type)+' hrs.'
        body_mail_html = '<p>Buen d&iacute;a.</p><p>En relaci&oacute;n al proceso de MC IMMEX, comparto el formato de solicitud para el proceso debido con categor&iacute;a de '+str(r_type)+' hrs.</p>'
        #ya tenemos todo mandemos el correo
        mail_pool = self.env['mail.mail']
        values={}
        values.update({'subject': self.name})
        values.update({'email_to': emto})
        values.update({'body_html': body_mail_html })
        values.update({'body': body_mail })
        values.update({'attachment_ids': inserted_id })
        values.update({'res_id': self.id }) #[optional] here is the record id, where you want to post that email after sending
        values.update({'model': 'project.task' }) #[optional] here is the object(like 'project.project')  to whose record id you want to post that email after sending
        msg_id = mail_pool.create(values)
        if msg_id:
            mail_pool.send([msg_id])
        self.email_sent = 1
        return {'type': 'ir.actions.act_url','name': filename,'url': url} #regresamos el link con el archivo         


        
        